"""
spreadsheet_converter.py
Handles:
  Excel  → CSV  (one CSV per sheet, or zipped if multiple sheets)
  CSV    → Excel (with auto-formatting, column widths, header styling)
"""

import io
import os
import zipfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Excel → CSV
# ---------------------------------------------------------------------------

def excel_to_csv(excel_path: str, output_path: str) -> tuple[str, bool]:
    """
    Convert Excel to CSV.

    If the workbook has ONE sheet  → writes a single .csv file.
    If the workbook has MANY sheets → writes a .zip containing one CSV per sheet.

    Returns (output_path, is_zip).
    """
    all_sheets = pd.read_excel(excel_path, sheet_name=None, dtype=str)
    sheet_names = list(all_sheets.keys())

    if len(sheet_names) == 1:
        df = all_sheets[sheet_names[0]]
        df.to_csv(output_path, index=False, encoding="utf-8-sig")   # utf-8-sig for Excel compatibility
        return output_path, False

    # Multiple sheets → zip
    zip_path = output_path.replace(".csv", ".zip") if output_path.endswith(".csv") else output_path + ".zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for sheet_name in sheet_names:
            df  = all_sheets[sheet_name]
            buf = io.StringIO()
            df.to_csv(buf, index=False, encoding="utf-8")
            safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in sheet_name)
            zf.writestr(f"{safe_name}.csv", buf.getvalue())

    return zip_path, True


# ---------------------------------------------------------------------------
# CSV → Excel
# ---------------------------------------------------------------------------

def csv_to_excel(csv_path: str, output_path: str) -> str:
    """
    Convert CSV to a styled Excel file.
    - Header row: bold white text on indigo background
    - Alternating row shading
    - Auto column widths
    - Freeze top row
    """
    df = pd.read_csv(csv_path, dtype=str).fillna("")

    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(csv_path))[0][:31]  # Excel sheet name ≤ 31 chars

    # --- Styles ---
    HEADER_FILL = PatternFill("solid", fgColor="F97316")   # orange-500
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ROW_FILL_ODD  = PatternFill("solid", fgColor="F8FAFC")
    ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
    ROW_FONT  = Font(name="Arial", size=10)
    ROW_ALIGN = Alignment(vertical="center", wrap_text=False)

    thin = Side(style="thin", color="E2E8F0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Write header ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    ws.row_dimensions[1].height = 28

    # --- Write data rows ---
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ROW_FILL_ODD if row_idx % 2 != 0 else ROW_FILL_EVEN
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else None)
            cell.fill      = fill
            cell.font      = ROW_FONT
            cell.alignment = ROW_ALIGN
            cell.border    = BORDER
        ws.row_dimensions[row_idx].height = 18

    # --- Auto column widths ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter  = get_column_letter(col_idx)
        header_len  = len(str(col_name)) + 2
        max_data_len = df.iloc[:, col_idx - 1].astype(str).map(len).max() if len(df) > 0 else 0
        col_width   = min(max(header_len, max_data_len + 2, 8), 50)
        ws.column_dimensions[col_letter].width = col_width

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    if len(df.columns) > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col}1"

    wb.save(output_path)
    return output_path
